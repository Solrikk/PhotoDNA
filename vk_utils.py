# vk_utils.py
import os
import time
import logging
import requests
from bs4 import BeautifulSoup
from PIL import Image
from io import BytesIO
import pandas as pd
from tkinter import messagebox
from image_matcher import ImageMatcher
from yandex_search import search_yandex_image
from browser_utils import setup_browser

WAIT_TIME = 20
RESULTS_FILE = 'results.xlsx'
SIMILARITY_THRESHOLD = 0.89
MAX_TOTAL_RECORDS = 2000

def extract_vk_album_photos(driver, album_url):
    try:
        driver.get(album_url)
        logging.info(f"Переход на страницу альбома ВКонтакте: {album_url}")
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        wait = WebDriverWait(driver, WAIT_TIME)
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "a[href*='/photo-']")))
        time.sleep(5)
        last_height = driver.execute_script("return document.body.scrollHeight")
        scroll_attempt = 0
        max_scroll_attempts = 50  # Increased to load more photos
        while scroll_attempt < max_scroll_attempts:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(3)  # Increased delay to allow more photos to load
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                scroll_attempt += 1
                logging.info(f"Попытка прокрутки {scroll_attempt}/{max_scroll_attempts} не удалась. Новый контент не найден.")
            else:
                last_height = new_height
                scroll_attempt = 0
                logging.info(f"Успешная прокрутка. Текущая высота: {new_height}")
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        photo_links = soup.find_all('a', href=lambda href: href and href.startswith('/photo-'))
        photo_urls = ["https://vk.com" + link['href'] for link in photo_links]
        photo_urls = list(dict.fromkeys(photo_urls))
        logging.info(f"Найдено {len(photo_urls)} фотографий в альбоме.")
        return photo_urls
    except Exception as e:
        from selenium.webdriver.common.by import By
        logging.error(f"Ошибка при извлечении фото из альбома ВКонтакте: {str(e)}")
        return []

def extract_vk_photo_url(driver, photo_url):
    try:
        driver.get(photo_url)
        logging.info(f"Переход на страницу фотографии ВКонтакте: {photo_url}")
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        wait = WebDriverWait(driver, WAIT_TIME)
        wait.until(EC.presence_of_element_located((By.ID, "pv_photo")))
        time.sleep(5)
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, 'html.parser')
        pv_photo_div = soup.find('div', id='pv_photo')
        if pv_photo_div:
            img_tag = pv_photo_div.find('img')
            if img_tag and img_tag.get('src'):
                img_url = img_tag['src']
                logging.info(f"Найден URL фотографии: {img_url}")
                return img_url
        logging.warning("Не удалось найти URL фотографии на странице.")
        return None
    except Exception as e:
        logging.error(f"Ошибка при извлечении URL фотографии ВКонтакте: {str(e)}")
        return None

def process_images(album_url):
    from selenium.webdriver.common.by import By
    if not (album_url.startswith("https://vk.com/album") or album_url.startswith("https://vk.com/photo")):
        logging.error("Некорректная ссылка. Пожалуйста, введите ссылку на альбом или фотографию ВКонтакте.")
        return
    matcher = ImageMatcher()
    driver = setup_browser()
    all_data = []
    record_counter = [0]
    try:
        if album_url.startswith("https://vk.com/album"):
            photo_urls = extract_vk_album_photos(driver, album_url)
            if not photo_urls:
                logging.info("Не удалось извлечь фотографии из альбома.")
                return
            for idx, photo_url in enumerate(photo_urls, 1):
                try:
                    logging.info(f"Обработка фотографии {idx}/{len(photo_urls)}: {photo_url}")
                    img_url = extract_vk_photo_url(driver, photo_url)
                    if not img_url:
                        logging.warning(f"Не удалось извлечь URL фотографии: {photo_url}. Пропуск.")
                        continue
                    response = requests.get(img_url, timeout=10)
                    img = Image.open(BytesIO(response.content)).convert('RGB')
                    temp_image_path = f"temp_image_{idx}.jpg"
                    img.save(temp_image_path)
                    source_embedding = matcher.get_embedding(temp_image_path)
                    if source_embedding is None:
                        logging.warning(f"Не удалось извлечь эмбеддинг для: {img_url}. Пропуск.")
                        os.remove(temp_image_path)
                        continue
                    logging.info(f"Извлечён эмбеддинг для: {img_url}")
                    search_yandex_image(driver, temp_image_path, matcher, source_embedding, f"Фото {idx}", all_data, record_counter, img_url)
                    os.remove(temp_image_path)
                    if record_counter[0] >= MAX_TOTAL_RECORDS:
                        logging.info(f"Достигнуто максимальное количество записей: {MAX_TOTAL_RECORDS}")
                        break
                    time.sleep(1)
                except Exception as e:
                    logging.error(f"Ошибка при обработке фотографии {photo_url}: {str(e)}")
                    continue
        elif album_url.startswith("https://vk.com/photo"):
            img_url = extract_vk_photo_url(driver, album_url)
            if not img_url:
                logging.info("Не удалось извлечь URL фотографии.")
                return
            try:
                logging.info(f"Обработка фотографии: {img_url}")
                response = requests.get(img_url, timeout=10)
                img = Image.open(BytesIO(response.content)).convert('RGB')
                temp_image_path = "temp_image_single.jpg"
                img.save(temp_image_path)
                source_embedding = matcher.get_embedding(temp_image_path)
                if source_embedding is None:
                    logging.warning(f"Не удалось извлечь эмбеддинг для: {img_url}. Пропуск.")
                    os.remove(temp_image_path)
                    return
                logging.info(f"Извлечён эмбеддинг для: {img_url}")
                search_yandex_image(driver, temp_image_path, matcher, source_embedding, "Фотография", all_data, record_counter, img_url)
                os.remove(temp_image_path)
            except Exception as e:
                logging.error(f"Ошибка при обработке фотографии {img_url}: {str(e)}")
        if all_data:
            df = pd.DataFrame(all_data)
            df = df[df['Схожесть'].astype(float) > SIMILARITY_THRESHOLD].reset_index(drop=True)
            for col in ['Дата публикации', 'Текст поста', 'Переслано от', 'Заголовок', 'IP адрес']:
                if col not in df.columns:
                    df[col] = ''
            excel_file = RESULTS_FILE
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Результаты')
                workbook = writer.book
                worksheet = writer.sheets['Результаты']
                from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
                from openpyxl.utils import get_column_letter
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))
                alignment_wrap = Alignment(wrap_text=True)
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    column_letter = get_column_letter(col)
                    max_length = max(df.iloc[:, col - 1].astype(str).map(len).max(), len(str(cell.value)))
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = alignment_wrap
                        if df.columns[col - 1] in ['URL сайта', 'URL исходного фото', 'URL найденного фото']:
                            url = cell.value
                            if pd.notna(url) and url:
                                cell.value = "Ссылка"
                                cell.hyperlink = url
                                cell.style = "Hyperlink"
                if 'Схожесть' in df.columns:
                    similarity_col = df.columns.get_loc('Схожесть') + 1
                    for row in range(2, len(df) + 2):
                        similarity_value = df.iloc[row - 2]['Схожесть']
                        try:
                            similarity_percent = float(similarity_value)
                        except:
                            similarity_percent = 0.0
                        if similarity_percent >= SIMILARITY_THRESHOLD:
                            color_intensity = min(int((similarity_percent - SIMILARITY_THRESHOLD) / (1.0 - SIMILARITY_THRESHOLD) * 255), 255)
                            fill_color = f"{255:02X}{255 - color_intensity:02X}{255 - color_intensity:02X}"
                            cell = worksheet.cell(row=row, column=similarity_col)
                            cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell = worksheet.cell(row=row, column=similarity_col)
                        cell.border = thin_border
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
            logging.info(f"Данные успешно записаны в {excel_file}")
            messagebox.showinfo("Завершено", f"Данные успешно записаны в {excel_file}")
        else:
            logging.info("Нет данных для записи.")
            messagebox.showinfo("Завершено", "Нет данных для записи.")
    except Exception as e:
        logging.error(f"Фатальная ошибка: {str(e)}")
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
    finally:
        driver.quit()
        logging.info("Браузер закрыт")
