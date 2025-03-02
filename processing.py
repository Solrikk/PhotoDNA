
import os
import logging
import requests
from PIL import Image
from io import BytesIO
import time
from datetime import datetime
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from concurrent.futures import ThreadPoolExecutor, as_completed
from tkinter import messagebox

from browser_utils import setup_browser
from image_matcher import ImageMatcher
from vk_utils import extract_vk_album_photos, extract_vk_photo_url
from search_utils import search_yandex_image

SIMILARITY_THRESHOLD = 0.89
RESULTS_FILE = 'results.xlsx'

def cleanup_temp_files():
    try:
        import glob
        for temp_file in glob.glob("temp_image_*.jpg"):
            try:
                os.remove(temp_file)
                logging.info(f"Удален временный файл: {temp_file}")
            except:
                pass
    except Exception as e:
        logging.error(f"Ошибка при очистке временных файлов: {str(e)}")


def process_images(album_url, vk_login, vk_password, show_completion=True):
    if not (album_url.startswith("https://vk.com/album")
            or album_url.startswith("https://vk.com/photo")):
        logging.error(
            "Некорректная ссылка. Пожалуйста, введите ссылку на альбом или фотографию ВКонтакте."
        )
        return

    def process_single_photo(photo_data):
        try:
            photo_url, idx, total = photo_data
            driver = setup_browser()
            matcher = ImageMatcher()
            try:
                logging.info(f"Обработка фотографии {idx}/{total}: {photo_url}")
                img_url = extract_vk_photo_url(driver, photo_url)
                if not img_url:
                    return None
                response = requests.get(img_url, timeout=5)
                img = Image.open(BytesIO(response.content)).convert('RGB')
                temp_image_path = f"temp_image_{idx}.jpg"
                img.save(temp_image_path)
                source_embedding = matcher.get_embedding(temp_image_path)
                if source_embedding is None:
                    os.remove(temp_image_path)
                    return None
                results = []
                search_yandex_image(driver, temp_image_path, matcher,
                                    source_embedding, f"Фото {idx}",
                                    results, [0], img_url)
                os.remove(temp_image_path)
                return results
            finally:
                driver.quit()
        except Exception as e:
            logging.error(f"Ошибка при обработке фото {photo_url}: {str(e)}")
            return None
            
    matcher = ImageMatcher()
    driver = setup_browser()
    all_data = []
    record_counter = [0]
    try:
        if album_url.startswith("https://vk.com/album"):
            driver = setup_browser()
            photo_urls = extract_vk_album_photos(driver, album_url)
            driver.quit()

            if not photo_urls:
                logging.info("Не удалось извлечь фотографии из альбома.")
                return

            all_data = []
            with ThreadPoolExecutor(max_workers=4) as executor:
                photo_data = [(url, idx, len(photo_urls))
                              for idx, url in enumerate(photo_urls, 1)]
                future_to_photo = {executor.submit(process_single_photo, data): data
                                   for data in photo_data}

                for future in as_completed(future_to_photo):
                    photo_data = future_to_photo[future]
                    try:
                        results = future.result()
                        if results:
                            all_data.extend(results)
                    except Exception as e:
                        logging.error(
                            f"Ошибка при обработке фото {photo_data[0]}: {str(e)}")
                        continue

        elif album_url.startswith("https://vk.com/photo"):
            img_url = extract_vk_photo_url(driver, album_url)
            if not img_url:
                logging.info("Не удалось извлечь URL фотографии.")
                return
            try:
                logging.info(f"Обработка фотографии: {img_url}")
                response = requests.get(img_url, timeout=5)
                img = Image.open(BytesIO(response.content)).convert('RGB')
                temp_image_path = "temp_image_single.jpg"
                img.save(temp_image_path)
                source_embedding = matcher.get_embedding(temp_image_path)
                if source_embedding is None:
                    logging.warning(
                        f"Не удалось извлечь эмбеддинг для: {img_url}. Пропуск."
                    )
                    os.remove(temp_image_path)
                    return
                logging.info(f"Извлечён эмбеддинг для: {img_url}")
                search_yandex_image(driver, temp_image_path, matcher,
                                    source_embedding, "Фотография", all_data,
                                    record_counter, img_url)
                os.remove(temp_image_path)
            except Exception as e:
                logging.error(
                    f"Ошибка при обработке фотографии {img_url}: {str(e)}")
                
        if all_data:
            df = pd.DataFrame(all_data)
            df = df[df['Схожесть'].astype(float) >
                    SIMILARITY_THRESHOLD].reset_index(drop=True)
            for col in [
                    'Дата публикации', 'Текст поста', 'Переслано от',
                    'Заголовок', 'IP адрес'
            ]:
                if col not in df.columns:
                    df[col] = ''
            excel_file = RESULTS_FILE
            with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Результаты')
                workbook = writer.book
                worksheet = writer.sheets['Результаты']
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD",
                                          end_color="4F81BD",
                                          fill_type="solid")
                thin_border = Border(left=Side(style='thin'),
                                     right=Side(style='thin'),
                                     top=Side(style='thin'),
                                     bottom=Side(style='thin'))
                alignment_wrap = Alignment(wrap_text=True)
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                    column_letter = get_column_letter(col)
                    max_length = max(
                        df.iloc[:, col - 1].astype(str).map(len).max(),
                        len(str(cell.value)))
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[
                        column_letter].width = adjusted_width
                    for row in range(2, len(df) + 2):
                        cell = worksheet.cell(row=row, column=col)
                        cell.alignment = alignment_wrap
                        if df.columns[col - 1] in [
                                'URL сайта', 'URL исходного фото',
                                'URL найденного фото'
                        ]:
                            url = cell.value
                            if pd.notna(url) and url:
                                cell.value = "Ссылка"
                                cell.hyperlink = url
                                cell.style = "Hyperlink"
                if 'Схожесть' in df.columns:
                    similarity_col = df.columns.get_loc('Схожесть') + 1
                    for row in range(2, len(df) + 2):
                        similarity_value = df.iloc[row - 2]['Схожесть']
                        try:
                            similarity_percent = float(similarity_value)
                        except:
                            similarity_percent = 0.0
                        if similarity_percent >= SIMILARITY_THRESHOLD:
                            color_intensity = min(
                                int((similarity_percent - SIMILARITY_THRESHOLD)
                                    / (1.0 - SIMILARITY_THRESHOLD) * 255), 255)
                            fill_color = f"{255:02X}{255 - color_intensity:02X}{255 - color_intensity:02X}"
                            cell = worksheet.cell(row=row,
                                                  column=similarity_col)
                            cell.fill = PatternFill(start_color=fill_color,
                                                    end_color=fill_color,
                                                    fill_type="solid")
                        cell = worksheet.cell(row=row, column=similarity_col)
                        cell.border = thin_border
                for row in range(2, len(df) + 2):
                    for col in range(1, len(df.columns) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
            if show_completion:
                logging.info(f"Данные успешно записаны в {excel_file}")
                messagebox.showinfo("Завершено",
                                    f"Данные успешно записаны в {excel_file}")
            else:
                logging.info("Обработка завершена без вывода сообщения.")
        else:
            if show_completion:
                logging.info("Нет данных для записи.")
                messagebox.showinfo("Завершено", "Нет данных для записи.")
    except Exception as e:
        logging.error(f"Фатальная ошибка: {str(e)}")
        messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
    finally:
        driver.quit()
        logging.info("Браузер закрыт")
