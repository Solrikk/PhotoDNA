        import os

        os.environ["KMP_DUPLICATE_LIB_OK"] = "TRUE"
        os.environ['TF_ENABLE_ONEDNN_OPTS'] = '0'
        os.environ['OMP_NUM_THREADS'] = '1'
        import time
        import logging
        import numpy as np
        import requests
        from PIL import Image
        from io import BytesIO
        import tensorflow as tf
        from tensorflow.keras.applications import ResNet50
        from tensorflow.keras.applications.resnet50 import preprocess_input
        from tensorflow.keras.preprocessing import image
        from scipy.spatial.distance import cosine
        from selenium import webdriver
        from selenium.webdriver.common.by import By
        from selenium.webdriver.support.ui import WebDriverWait
        from selenium.webdriver.support import expected_conditions as EC
        from selenium.webdriver.edge.service import Service as EdgeService
        from selenium.webdriver.edge.options import Options as EdgeOptions
        from webdriver_manager.microsoft import EdgeChromiumDriverManager
        from bs4 import BeautifulSoup
        import spacy
        from datetime import datetime
        import dateparser
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.hyperlink import Hyperlink
        from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException
        from selenium.webdriver.common.action_chains import ActionChains
        import socket
        from sklearn.feature_extraction.text import TfidfVectorizer
        import tkinter as tk
        from tkinter import messagebox, filedialog, ttk
        import multiprocessing

        WAIT_TIME = 10
        RESULTS_FILE = 'results.xlsx'
        SIMILARITY_THRESHOLD = 0.89
        MAX_ADDITIONAL_IMAGES = 2
        MAX_TOTAL_RECORDS = 2000
        MAX_ADDITIONAL_SEARCHES_PER_IMAGE = 2

        NUM_PARALLEL_PROCESSES = 4

        logging.basicConfig(level=logging.INFO,
                            format='%(asctime)s - %(levelname)s - %(message)s')


        class ImageMatcher:

            def __init__(self):
                logging.info("Загрузка модели ResNet50...")
                self.model = ResNet50(weights='imagenet',
                                      include_top=False,
                                      pooling='avg')
                logging.info("Загрузка модели spaCy...")
                try:
                    self.nlp = spacy.load("ru_core_news_sm")
                except OSError:
                    logging.info(
                        "Модель spaCy 'ru_core_news_sm' не найдена. Загружаю...")
                    from spacy.cli import download
                    download("ru_core_news_sm")
                    self.nlp = spacy.load("ru_core_news_sm")
                logging.info("Инициализация TF-IDF векторизатора...")
                self.vectorizer = TfidfVectorizer(max_features=10,
                                                  stop_words='russian')

            def preprocess_image(self, img_path):
                try:
                    if isinstance(img_path, str):
                        img = image.load_img(img_path, target_size=(224, 224))
                    else:
                        img = img_path.resize((224, 224), Image.Resampling.LANCZOS)
                    x = np.asarray(img, dtype=np.float32)
                    x = np.expand_dims(x, axis=0)
                    return preprocess_input(x)
                except Exception as e:
                    logging.warning(
                        f"Ошибка при предварительной обработке изображения: {str(e)}")
                    return None

            def get_embedding(self, img):
                try:
                    preprocessed = self.preprocess_image(img)
                    if preprocessed is not None:
                        return self.model.predict(preprocessed, verbose=0)
                    return None
                except Exception as e:
                    logging.warning(f"Ошибка при получении эмбеддинга: {str(e)}")
                    return None

            @staticmethod
            def compare_images_batch(embedding1, embeddings_list):
                try:
                    similarities = 1 - np.array([
                        cosine(embedding1.flatten(), emb.flatten())
                        for emb in embeddings_list
                    ])
                    return similarities
                except Exception as e:
                    logging.warning(
                        f"Ошибка при пакетном сравнении изображений: {str(e)}")
                    return np.zeros(len(embeddings_list))

            def compare_images(self, embedding1, embedding2):
                try:
                    return float(1 -
                                 cosine(embedding1.flatten(), embedding2.flatten()))
                except Exception as e:
                    logging.warning(f"Ошибка при сравнении изображений: {str(e)}")
                    return 0.0

            def download_and_process_image(self, url):
                try:
                    response = requests.get(url, timeout=10)
                    img = Image.open(BytesIO(response.content)).convert('RGB')
                    return self.get_embedding(img)
                except Exception as e:
                    logging.warning(f"Ошибка при обработке URL {url}: {str(e)}")
                    return None

            def fetch_page_text(self, url):
                try:
                    response = requests.get(url, timeout=10)
                    soup = BeautifulSoup(response.content, 'html.parser')
                    texts = soup.stripped_strings
                    return ' '.join(texts)
                except Exception as e:
                    logging.warning(f"Ошибка при получении страницы {url}: {str(e)}")
                    return ""

            def generate_title(self, text):
                try:
                    if '—' in text:
                        title = text.split('—')[0].strip()
                    else:
                        title = text.strip()
                    if not title:
                        logging.warning("Заголовок пуст после извлечения.")
                        return None
                    logging.info(f"Извлечённый заголовок: {title}")
                    return title
                except Exception as e:
                    logging.warning(f"Ошибка при генерации заголовка: {str(e)}")
                    return None

            def extract_image_info(self, soup, image_url):
                try:
                    img_tags = soup.find_all('img', src=image_url)
                    for img in img_tags:
                        title = img.get('title')
                        alt = img.get('alt')
                        if title:
                            return title
                        if alt:
                            return alt
                    return None
                except:
                    return None

            def extract_publication_date(self, soup):
                try:
                    meta_date = soup.find('meta',
                                          attrs={'property': 'article:published_time'})
                    if meta_date and meta_date.get('content'):
                        return meta_date.get('content')
                    for tag in soup.find_all(['span', 'div', 'p']):
                        text = tag.get_text(strip=True)
                        date = dateparser.parse(text, languages=['ru'])
                        if date:
                            return date.isoformat()
                    return None
                except Exception as e:
                    logging.warning(f"Ошибка при извлечении даты публикации: {str(e)}")
                    return None


        def setup_browser():
            options = EdgeOptions()
            options.add_argument('--start-maximized')
            options.add_argument('--no-sandbox')
            options.add_argument('--disable-dev-shm-usage')
            options.add_argument('--disable-blink-features=AutomationControlled')
            options.add_experimental_option('excludeSwitches', ['enable-automation'])
            options.add_experimental_option('useAutomationExtension', False)
            options.add_argument('--disable-gpu')
            options.add_argument('--disable-software-rasterizer')

            try:
                driver_manager = EdgeChromiumDriverManager()
                driver_path = driver_manager.install()
                service = EdgeService(driver_path)
                driver = webdriver.Edge(service=service, options=options)
                logging.info(f"Edge драйвер успешно установлен: {driver_path}")
            except Exception as e:
                logging.error(f"Ошибка при инициализации драйвера: {str(e)}")
                raise e

            driver.execute_script(
                "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
            )
            return driver


        def click_element_js(driver, element):
            try:
                driver.execute_script("arguments[0].scrollIntoView(true);", element)
                ActionChains(driver).move_to_element(element).perform()
                driver.execute_script("arguments[0].click();", element)
            except Exception as e:
                logging.warning(f"Ошибка при клике на элемент: {str(e)}")


        def get_ip_address(url):
            try:
                hostname = requests.utils.urlparse(url).hostname
                ip = socket.gethostbyname(hostname)
                return ip
            except:
                return None


        def perform_additional_search(driver, query, matcher, data_list,
                                      record_counter, processed_ips, source_embedding):
            if not query:
                logging.warning("Пустой запрос. Пропуск дополнительного поиска.")
                return
            if len(query) > 100:
                logging.warning(f"Запрос слишком длинный: {query[:100]}... Пропуск.")
                return
            try:
                driver.get('https://yandex.ru/')
                wait = WebDriverWait(driver, WAIT_TIME)
                search_input = wait.until(
                    EC.presence_of_element_located((By.NAME, "text")))
                search_input.clear()
                search_input.send_keys(query)
                search_input.submit()
                logging.info(f"Выполнен дополнительный поиск по запросу: {query}")
                driver.save_screenshot(f'yandex_additional_search_{query[:50]}.png')
                time.sleep(2)
                items = wait.until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "li.serp-item")))
                logging.info(
                    f"Найдено {len(items)} результатов для дополнительного поиска по запросу: {query}"
                )
                for item in items:
                    try:
                        link_element = item.find_element(By.CSS_SELECTOR, 'a.Link')
                        site_link = link_element.get_attribute('href')
                        if site_link:
                            ip_address = get_ip_address(site_link)
                            if ip_address and ip_address not in processed_ips:
                                processed_ips.add(ip_address)
                                response = requests.get(site_link, timeout=10)
                                soup = BeautifulSoup(response.content, 'html.parser')
                                img_tag = soup.find('img')
                                if img_tag and img_tag.get('src'):
                                    img_url = requests.compat.urljoin(
                                        site_link, img_tag.get('src'))
                                    comparison_embedding = matcher.download_and_process_image(
                                        img_url)
                                    if comparison_embedding is not None:
                                        similarity = matcher.compare_images(
                                            source_embedding, comparison_embedding)
                                        if similarity > SIMILARITY_THRESHOLD:
                                            row_data = {
                                                'Запрос': query,
                                                'URL сайта': site_link,
                                                'IP адрес': ip_address,
                                                'Схожесть': f"{similarity:.4f}",
                                                'URL исходного фото': '',
                                                'URL найденного фото': img_url
                                            }
                                            data_list.append(row_data)
                                            record_counter[0] += 1
                                            logging.info(
                                                f"Добавлены данные из дополнительного поиска: {site_link} (Схожесть: {similarity:.4f})"
                                            )
                                            if record_counter[0] >= MAX_TOTAL_RECORDS:
                                                return
                    except Exception as e:
                        logging.warning(
                            f"Ошибка при обработке дополнительного результата: {str(e)}"
                        )
                        continue
            except Exception as e:
                logging.error(
                    f"Ошибка при выполнении дополнительного поиска для запроса '{query}': {str(e)}"
                )


        def collect_similar_images(driver, wait, matcher, source_embedding, image_name,
                                   data_list, record_counter, original_photo_url):
            try:
                try:
                    time.sleep(5)
                    more_button = wait.until(
                        EC.element_to_be_clickable(
                            (By.CLASS_NAME, "CbirSites-MoreButton")))
                    click_element_js(driver, more_button)
                    logging.info("Нажата кнопка 'Показать все' на Яндексе.")
                    driver.save_screenshot('yandex_step_clicked_more.png')
                    time.sleep(5)
                except Exception as e:
                    logging.error(f"Не удалось нажать кнопку 'Показать все': {str(e)}")
                last_height = driver.execute_script(
                    "return document.body.scrollHeight")
                scroll_attempt = 0
                max_scroll_attempts = 15
                while scroll_attempt < max_scroll_attempts:
                    current_height = driver.execute_script(
                        "return window.pageYOffset;")
                    scroll_height = driver.execute_script(
                        "return document.body.scrollHeight;")
                    step = (scroll_height - current_height) / 4
                    for i in range(4):
                        driver.execute_script(
                            f"window.scrollTo(0, {current_height + step * (i+1)});")
                        time.sleep(1)
                    logging.info(f"Прокрутка вниз... Попытка {scroll_attempt + 1}")
                    time.sleep(3)
                    new_height = driver.execute_script(
                        "return document.body.scrollHeight")
                    if new_height == last_height:
                        scroll_attempt += 1
                        logging.info("Новый контент не загружен.")
                    else:
                        last_height = new_height
                        scroll_attempt = 0
                logging.info("Прокрутка завершена.")
                items = wait.until(
                    EC.presence_of_all_elements_located(
                        (By.CSS_SELECTOR, "li.CbirSites-Item")))
                logging.info(
                    f"Найдено {len(items)} элементов с классом 'CbirSites-Item'")
                processed_ips = set()
                for idx, item in enumerate(items):
                    try:
                        img_element = item.find_element(By.TAG_NAME, 'img')
                        thumbnail_url = img_element.get_attribute('src')
                        site_link_element = item.find_element(
                            By.CSS_SELECTOR, 'a.Link_view_outer.CbirSites-ItemDomain')
                        site_link = site_link_element.get_attribute('href')
                        logging.info(f"Обработка элемента {idx + 1}/{len(items)}")
                        logging.info(f"URL миниатюры: {thumbnail_url}")
                        logging.info(f"URL сайта: {site_link}")
                        similarity = 0.0
                        row_data = {
                            'Источник': 'Яндекс',
                            'Название изображения': image_name,
                            'URL миниатюры': thumbnail_url,
                            'URL сайта': site_link,
                            'Схожесть': '',
                            'URL исходного фото': original_photo_url,
                            'URL найденного фото': ''
                        }
                        if thumbnail_url and site_link:
                            comparison_embedding = matcher.download_and_process_image(
                                thumbnail_url)
                            if comparison_embedding is not None:
                                similarity = matcher.compare_images(
                                    source_embedding, comparison_embedding)
                                row_data['Схожесть'] = f"{similarity:.4f}"
                                logging.info(f"Оценка схожости: {similarity}")
                        if similarity > SIMILARITY_THRESHOLD:
                            try:
                                driver.execute_script("window.open('');")
                                driver.switch_to.window(driver.window_handles[1])
                                driver.get(site_link)
                                wait_short = WebDriverWait(driver, WAIT_TIME)
                                wait_short.until(
                                    EC.presence_of_element_located(
                                        (By.TAG_NAME, 'body')))
                                time.sleep(5)
                                page_content = driver.page_source
                                soup = BeautifulSoup(page_content, 'html.parser')
                                publication_date = matcher.extract_publication_date(
                                    soup)
                                row_data[
                                    'Дата публикации'] = publication_date if publication_date else ''
                                post_text = ''
                                forward_from = ''
                                post_body = soup.find('div', class_='post-body')
                                if post_body:
                                    post_text_element = post_body.find(
                                        'div', class_='post-text')
                                    if post_text_element:
                                        post_text = post_text_element.get_text(
                                            separator=' ', strip=True)
                                    forward_info = post_body.find('div',
                                                                  class_='post-from')
                                    if forward_info:
                                        forward_from = forward_info.get_text(
                                            strip=True).replace('Forward from: ', '')
                                row_data['Текст поста'] = post_text
                                row_data['Переслано от'] = forward_from
                                title = matcher.generate_title(post_text)
                                row_data['Заголовок'] = title if title else ''
                                ip_address = get_ip_address(site_link)
                                row_data['IP адрес'] = ip_address if ip_address else ''
                                row_data['URL найденного фото'] = thumbnail_url
                                data_list.append(row_data)
                                record_counter[0] += 1
                                logging.info(
                                    f"Добавлены данные для: {image_name} (Всего записей: {record_counter[0]})"
                                )
                                search_queries = []
                                if title:
                                    search_queries.append(f'"{title}"')
                                if post_text:
                                    search_queries.append(post_text)
                                if forward_from:
                                    search_queries.append(forward_from)
                                additional_searches_done = 0
                                for query in search_queries:
                                    if record_counter[0] >= MAX_TOTAL_RECORDS:
                                        break
                                    if additional_searches_done >= MAX_ADDITIONAL_SEARCHES_PER_IMAGE:
                                        break
                                    perform_additional_search(driver, query, matcher,
                                                              data_list,
                                                              record_counter,
                                                              processed_ips,
                                                              source_embedding)
                                    additional_searches_done += 1
                                if record_counter[0] >= MAX_TOTAL_RECORDS:
                                    logging.info(
                                        f"Достигнуто максимальное количество записей: {MAX_TOTAL_RECORDS}"
                                    )
                                    break
                            except Exception as e:
                                logging.error(
                                    f"Ошибка при обработке ссылки с высокой схожестью {site_link}: {str(e)}"
                                )
                            finally:
                                driver.close()
                                driver.switch_to.window(driver.window_handles[0])
                    except Exception as e:
                        logging.error(
                            f"Ошибка при обработке элемента {idx + 1}: {str(e)}")
                        continue
            except Exception as e:
                logging.error(f"Ошибка при сборе изображений: {str(e)}")


        def search_yandex_image(driver, image_path, matcher, source_embedding,
                                image_name, data_list, record_counter,
                                original_photo_url):
            try:
                driver.get('https://yandex.ru/images/')
                logging.info("Переход на Яндекс Картинки.")
                driver.save_screenshot('yandex_step_navigate.png')
                wait = WebDriverWait(driver, WAIT_TIME)
                try:
                    consent_button = wait.until(
                        EC.element_to_be_clickable(
                            (By.XPATH, "//button[contains(text(), 'Согласен')]")))
                    consent_button.click()
                    logging.info("Закрыто окно согласия на Яндексе.")
                    driver.save_screenshot('yandex_step_closed_consent.png')
                except:
                    logging.info("Окно согласия на Яндексе не найдено.")
                    driver.save_screenshot('yandex_step_no_consent.png')
                try:
                    upload_input = wait.until(
                        EC.presence_of_element_located(
                            (By.XPATH, "//input[@type='file']")))
                    upload_input.send_keys(os.path.abspath(image_path))
                    logging.info(f"Изображение загружено на Яндекс: {image_path}")
                    driver.save_screenshot('yandex_step_uploaded_image.png')
                except Exception as e:
                    logging.error(
                        "Не удалось найти или взаимодействовать с полем загрузки на Яндексе."
                    )
                    driver.save_screenshot('yandex_step_upload_error.png')
                    raise e
                time.sleep(2)
                try:
                    popup_close = wait.until(
                        EC.element_to_be_clickable(
                            (By.CLASS_NAME, "NeuroOnboarding-Close")))
                    popup_close.click()
                    logging.info("Закрыто всплывающее окно на Яндексе.")
                    driver.save_screenshot('yandex_step_closed_popup.png')
                except:
                    logging.info(
                        "Всплывающее окно на Яндексе не найдено или уже закрыто.")
                    driver.save_screenshot('yandex_step_no_popup.png')
                collect_similar_images(driver, wait, matcher, source_embedding,
                                       image_name, data_list, record_counter,
                                       original_photo_url)
            except Exception as e:
                logging.error(f"Ошибка при поиске на Яндексе: {str(e)}")
                driver.save_screenshot('yandex_step_search_error.png')


        def vk_login(driver):
            try:
                wait = WebDriverWait(driver, WAIT_TIME)
                #####################################
                # Ждем кнопку "Войти другим способом"
                #####################################
                other_login_btn = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//button[contains(., 'Войти другим способом')]")))
                other_login_btn.click()
                #####################################
                # Ввод номера телефона
                #####################################
                phone_input = wait.until(
                    EC.presence_of_element_located((By.NAME, "login")))
                phone_input.clear()
                phone_input.send_keys("+79275088557")
                #####################################
                # Ждем и вводим пароль
                #####################################
                password_input = wait.until(
                    EC.presence_of_element_located((By.NAME, "password")))
                password_input.send_keys("straleglans-qwE1solrikksolrikk")
                password_input.submit()

                time.sleep(2)

            except Exception as e:
                logging.error(f"Ошибка при авторизации ВКонтакте: {str(e)}")
                raise

        def extract_vk_album_photos(driver, album_url):
            try:
                driver.get(album_url)
                logging.info(f"Переход на страницу альбома ВКонтакте: {album_url}")
                #####################################
                # Проверяем необходимость авторизации
                #####################################
                try:
                    login_button = driver.find_element(By.XPATH, "//button[contains(., 'Войти другим способом')]")
                    if login_button:
                        logging.info("Требуется авторизация ВКонтакте")
                        vk_login(driver)
                        #####################################
                        # Возвращаемся на страницу альбома после авторизации
                        #####################################
                        driver.get(album_url)
                except:
                    logging.info("Авторизация не требуется")

                wait = WebDriverWait(driver, WAIT_TIME)
                time.sleep(10)
                wait.until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, "a[href*='/photo-']")))
                time.sleep(5)
                last_height = driver.execute_script(
                    "return document.body.scrollHeight")
                scroll_attempt = 0
                #####################################
                max_scroll_attempts = 8  # Уменьшено количество попыток прокрутки
                #####################################
                while scroll_attempt < max_scroll_attempts:
                    driver.execute_script(
                        "window.scrollTo(0, document.body.scrollHeight);")
                    time.sleep(3)  # Increased delay to allow more photos to load
                    new_height = driver.execute_script(
                        "return document.body.scrollHeight")
                    if new_height == last_height:
                        scroll_attempt += 1
                        logging.info(
                            f"Попытка прокрутки {scroll_attempt}/{max_scroll_attempts} не удалась. Новый контент не найден."
                        )
                    else:
                        last_height = new_height
                        scroll_attempt = 0
                        logging.info(
                            f"Успешная прокрутка. Текущая высота: {new_height}")
                page_source = driver.page_source
                soup = BeautifulSoup(page_source, 'html.parser')
                photo_links = soup.find_all(
                    'a', href=lambda href: href and href.startswith('/photo-'))
                photo_urls = ["https://vk.com" + link['href'] for link in photo_links]
                photo_urls = list(dict.fromkeys(photo_urls))
                logging.info(f"Найдено {len(photo_urls)} фотографий в альбоме.")
                return photo_urls
            except Exception as e:
                logging.error(
                    f"Ошибка при извлечении фото из альбома ВКонтакте: {str(e)}")
                return []


        def extract_vk_photo_url(driver, photo_url):
            try:
                driver.get(photo_url)
                logging.info(f"Переход на страницу фотографии ВКонтакте: {photo_url}")
                wait = WebDriverWait(driver, WAIT_TIME)
                wait.until(EC.presence_of_element_located((By.ID, "pv_photo")))
                time.sleep(5)
                page_source = driver.page_source
                soup = BeautifulSoup(page_source, 'html.parser')
                pv_photo_div = soup.find('div', id='pv_photo')
                if pv_photo_div:
                    img_tag = pv_photo_div.find('img')
                    if img_tag and img_tag.get('src'):
                        img_url = img_tag['src']
                        logging.info(f"Найден URL фотографии: {img_url}")
                        return img_url
                logging.warning("Не удалось найти URL фотографии на странице.")
                return None
            except Exception as e:
                logging.error(
                    f"Ошибка при извлечении URL фотографии ВКонтакте: {str(e)}")
                return None


        def process_images(album_url):
            if not (album_url.startswith("https://vk.com/album")
                    or album_url.startswith("https://vk.com/photo")):
                logging.error(
                    "Некорректная ссылка. Пожалуйста, введите ссылку на альбом или фотографию ВКонтакте."
                )
                return
            matcher = ImageMatcher()
            driver = setup_browser()
            all_data = []
            record_counter = [0]
            try:
                if album_url.startswith("https://vk.com/album"):
                    photo_urls = extract_vk_album_photos(driver, album_url)
                    if not photo_urls:
                        logging.info("Не удалось извлечь фотографии из альбома.")
                        return
                    for idx, photo_url in enumerate(photo_urls, 1):
                        try:
                            logging.info(
                                f"Обработка фотографии {idx}/{len(photo_urls)}: {photo_url}"
                            )
                            img_url = extract_vk_photo_url(driver, photo_url)
                            if not img_url:
                                logging.warning(
                                    f"Не удалось извлечь URL фотографии: {photo_url}. Пропуск."
                                )
                                continue
                            response = requests.get(img_url, timeout=10)
                            img = Image.open(BytesIO(response.content)).convert('RGB')
                            temp_image_path = f"temp_image_{idx}.jpg"
                            img.save(temp_image_path)
                            source_embedding = matcher.get_embedding(temp_image_path)
                            if source_embedding is None:
                                logging.warning(
                                    f"Не удалось извлечь эмбеддинг для: {img_url}. Пропуск."
                                )
                                os.remove(temp_image_path)
                                continue
                            logging.info(f"Извлечён эмбеддинг для: {img_url}")
                            search_yandex_image(driver, temp_image_path, matcher,
                                                source_embedding, f"Фото {idx}",
                                                all_data, record_counter, img_url)
                            os.remove(temp_image_path)
                            if record_counter[0] >= MAX_TOTAL_RECORDS:
                                logging.info(
                                    f"Достигнуто максимальное количество записей: {MAX_TOTAL_RECORDS}"
                                )
                                break
                            time.sleep(1)
                        except Exception as e:
                            logging.error(
                                f"Ошибка при обработке фотографии {photo_url}: {str(e)}"
                            )
                            continue
                elif album_url.startswith("https://vk.com/photo"):
                    img_url = extract_vk_photo_url(driver, album_url)
                    if not img_url:
                        logging.info("Не удалось извлечь URL фотографии.")
                        return
                    try:
                        logging.info(f"Обработка фотографии: {img_url}")
                        response = requests.get(img_url, timeout=10)
                        img = Image.open(BytesIO(response.content)).convert('RGB')
                        temp_image_path = "temp_image_single.jpg"
                        img.save(temp_image_path)
                        source_embedding = matcher.get_embedding(temp_image_path)
                        if source_embedding is None:
                            logging.warning(
                                f"Не удалось извлечь эмбеддинг для: {img_url}. Пропуск."
                            )
                            os.remove(temp_image_path)
                            return
                        logging.info(f"Извлечён эмбеддинг для: {img_url}")
                        search_yandex_image(driver, temp_image_path, matcher,
                                            source_embedding, "Фотография", all_data,
                                            record_counter, img_url)
                        os.remove(temp_image_path)
                    except Exception as e:
                        logging.error(
                            f"Ошибка при обработке фотографии {img_url}: {str(e)}")
                if all_data:
                    df = pd.DataFrame(all_data)
                    df = df[df['Схожесть'].astype(float) >
                            SIMILARITY_THRESHOLD].reset_index(drop=True)
                    for col in [
                            'Дата публикации', 'Текст поста', 'Переслано от',
                            'Заголовок', 'IP адрес'
                    ]:
                        if col not in df.columns:
                            df[col] = ''
                    excel_file = RESULTS_FILE
                    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
                        df.to_excel(writer, index=False, sheet_name='Результаты')
                        workbook = writer.book
                        worksheet = writer.sheets['Результаты']
                        header_font = Font(bold=True, color="FFFFFF")
                        header_fill = PatternFill(start_color="4F81BD",
                                                  end_color="4F81BD",
                                                  fill_type="solid")
                        thin_border = Border(left=Side(style='thin'),
                                             right=Side(style='thin'),
                                             top=Side(style='thin'),
                                             bottom=Side(style='thin'))
                        alignment_wrap = Alignment(wrap_text=True)
                        for col in range(1, len(df.columns) + 1):
                            cell = worksheet.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.border = thin_border
                            column_letter = get_column_letter(col)
                            max_length = max(
                                df.iloc[:, col - 1].astype(str).map(len).max(),
                                len(str(cell.value)))
                            adjusted_width = min(max_length + 2, 50)
                            worksheet.column_dimensions[
                                column_letter].width = adjusted_width
                            for row in range(2, len(df) + 2):
                                cell = worksheet.cell(row=row, column=col)
                                cell.alignment = alignment_wrap
                                if df.columns[col - 1] in [
                                        'URL сайта', 'URL исходного фото',
                                        'URL найденного фото'
                                ]:
                                    url = cell.value
                                    if pd.notna(url) and url:
                                        cell.value = "Ссылка"
                                        cell.hyperlink = url
                                        cell.style = "Hyperlink"
                        if 'Схожесть' in df.columns:
                            similarity_col = df.columns.get_loc('Схожесть') + 1
                            for row in range(2, len(df) + 2):
                                similarity_value = df.iloc[row - 2]['Схожесть']
                                try:
                                    similarity_percent = float(similarity_value)
                                except:
                                    similarity_percent = 0.0
                                if similarity_percent >= SIMILARITY_THRESHOLD:
                                    color_intensity = min(
                                        int((similarity_percent - SIMILARITY_THRESHOLD)
                                            / (1.0 - SIMILARITY_THRESHOLD) * 255), 255)
                                    fill_color = f"{255:02X}{255 - color_intensity:02X}{255 - color_intensity:02X}"
                                    cell = worksheet.cell(row=row,
                                                          column=similarity_col)
                                    cell.fill = PatternFill(start_color=fill_color,
                                                            end_color=fill_color,
                                                            fill_type="solid")
                                cell = worksheet.cell(row=row, column=similarity_col)
                                cell.border = thin_border
                        for row in range(2, len(df) + 2):
                            for col in range(1, len(df.columns) + 1):
                                cell = worksheet.cell(row=row, column=col)
                                cell.border = thin_border
                    logging.info(f"Данные успешно записаны в {excel_file}")
                    messagebox.showinfo("Завершено",
                                        f"Данные успешно записаны в {excel_file}")
                else:
                    logging.info("Нет данных для записи.")
                    messagebox.showinfo("Завершено", "Нет данных для записи.")
            except Exception as e:
                logging.error(f"Фатальная ошибка: {str(e)}")
                messagebox.showerror("Ошибка", f"Произошла ошибка: {str(e)}")
            finally:
                driver.quit()
                logging.info("Браузер закрыт")


        def start_processing(album_url):
            if not album_url:
                messagebox.showwarning(
                    "Внимание",
                    "Пожалуйста, введите ссылку на альбом или фотографию ВКонтакте.")
                return

            if not (album_url.startswith('https://vk.com/album') or album_url.startswith('https://vk.com/photo')):
                messagebox.showwarning(
                    "Ошибка",
                    "Неверный формат ссылки. Используйте ссылку на альбом или фотографию ВКонтакте.")
                return

            try:
                process_images(album_url)
            except Exception as e:
                logging.error(f"Критическая ошибка при обработке: {str(e)}")
                messagebox.showerror("Ошибка", "Произошла ошибка при обработке изображений")
                cleanup_temp_files()


        def create_gui():
            root = tk.Tk()
            root.title("PhotoDNA - Анализ схожих изображений")
            root.geometry("1000x600")
            root.configure(bg="#1e1e2e")
            root.resizable(False, False)

            style = ttk.Style()
            style.theme_use('default')
            style.configure("TLabel", 
                           font=("Montserrat", 12), 
                           background="#1e1e2e",
                           foreground="#ffffff")
            style.configure("TEntry", 
                           font=("Montserrat", 11),
                           fieldbackground="#2d2d3f",
                           foreground="#ffffff",
                           insertcolor="#ffffff")
            style.configure("TButton",
                           font=("Montserrat", 11, "bold"),
                           padding=12,
                           background="#7c3aed",
                           foreground="#ffffff")
            style.configure("Header.TLabel",
                           font=("Montserrat", 28, "bold"),
                           foreground="#ffffff",
                           background="#1e1e2e")
            style.configure("Main.TFrame", 
                           background="#1e1e2e")
            style.configure("Blue.Horizontal.TProgressbar",
                           background="#7c3aed",
                           troughcolor="#2d2d3f")

            main_frame = ttk.Frame(root, padding=30, style="Main.TFrame")
            main_frame.pack(fill=tk.BOTH, expand=True)

            header = ttk.Label(main_frame, text="PhotoDNA", style="Header.TLabel")
            header.pack(pady=(0, 5))

            subheader = ttk.Label(main_frame,
                                  text="Умный поиск и анализ изображений",
                                  font=("Segoe UI", 12),
                                  background="#f0f2f5",
                                  foreground="#666666")
            subheader.pack(pady=(0, 20))

            input_frame = ttk.Frame(main_frame, style="Main.TFrame")
            input_frame.pack(fill=tk.X, pady=10)

            label = ttk.Label(input_frame,
                              text="Введите ссылку на альбом или фотографию:",
                              font=("Segoe UI", 11))
            label.pack(pady=(0, 8))

            entry_style = ttk.Style()
            entry_style.configure("Custom.TEntry",
                                 fieldbackground="#2d2d3f",
                                 foreground="#ffffff",
                                 insertcolor="#ffffff",
                                 borderwidth=0)

            entry_frame = ttk.Frame(input_frame, style="Main.TFrame")
            entry_frame.pack(pady=(0, 5))

            entry = ttk.Entry(entry_frame, 
                             width=80, 
                             font=("Montserrat", 11),
                             style="Custom.TEntry")
            entry.pack(pady=(0, 5), ipady=8)

            def show_context_menu(event):
                context_menu.tk_popup(event.x_root, event.y_root)

            context_menu = tk.Menu(root, tearoff=0, bg="#2d2d3f", fg="#ffffff")
            context_menu.add_command(label="Вставить",
                                   command=lambda: entry.event_generate("<<Paste>>"),
                                   activebackground="#7c3aed",
                                   activeforeground="#ffffff")
            context_menu.add_command(label="Копировать",
                                   command=lambda: entry.event_generate("<<Copy>>"),
                                   activebackground="#7c3aed",
                                   activeforeground="#ffffff")
            entry.bind("<Button-3>", show_context_menu)

            progress_frame = ttk.Frame(main_frame, style="Main.TFrame")
            progress_frame.pack(fill=tk.X, pady=15)

            progress = ttk.Progressbar(progress_frame,
                                       orient=tk.HORIZONTAL,
                                       length=700,
                                       mode='indeterminate',
                                       style="Blue.Horizontal.TProgressbar")
            progress.pack()

            def on_start():
                album_url = entry.get().strip()
                if album_url:
                    start_button.configure(state='disabled')
                    progress.start()
                    root.update_idletasks()
                    start_processing(album_url)
                    progress.stop()
                    start_button.configure(state='normal')
                else:
                    messagebox.showwarning(
                        "PhotoDNA",
                        "Пожалуйста, введите ссылку на альбом или фотографию.")

            def paste(event):
                try:
                    entry.event_generate("<<Paste>>")
                except:
                    pass

            entry.bind("<Control-v>", paste)
            entry.bind("<Command-v>", paste)

            start_button = ttk.Button(main_frame,
                                      text="Начать обработку",
                                      command=on_start)
            start_button.pack(pady=20)

            root.mainloop()


        if __name__ == "__main__":
            create_gui()
        def cleanup_temp_files():
            try:
                import glob
                for temp_file in glob.glob("temp_image_*.jpg"):
                    try:
                        os.remove(temp_file)
                        logging.info(f"Удален временный файл: {temp_file}")
                    except:
                        pass
            except Exception as e:
                logging.error(f"Ошибка при очистке временных файлов: {str(e)}")